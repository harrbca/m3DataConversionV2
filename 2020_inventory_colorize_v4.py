import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from path_manager import PathManager
from collections import defaultdict

def autosize_and_shade_roll_groups(file_path):
    fill_gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    fill_red  = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # === STEP 1: Header → column‐index map (uppercase keys) ===
        headers = {
            str(cell.value).strip().upper(): idx + 1
            for idx, cell in enumerate(ws[1])
        }

        # Required for shading logic and cost/quantity:
        roll_col_index  = headers.get("RROLL#")
        cost_col_index  = headers.get("RLASTC")
        ware_col_index  = headers.get("RWARE#")
        item_col_index  = headers.get("ITEMNUMBER") or headers.get("ITEM#")
        rlrctd_index    = headers.get("RLRCTD")
        onhand_index    = headers.get("RONHAN")

        if not all([ware_col_index, item_col_index, cost_col_index, rlrctd_index, onhand_index]):
            print(f"⚠️ Skipping sheet '{sheet_name}': missing one of RWARE#, ITEMNUMBER/ITEM#, RLASTC, RLRCTD, or RONHAN.")
            continue

        # === STEP 2: Build an initial group_map by (warehouse, item, roll) ===
        # (Used to decide LOT# for each row.)
        initial_group = defaultdict(list)
        for row in range(2, ws.max_row + 1):
            warehouse  = ws.cell(row=row, column=ware_col_index).value
            item       = ws.cell(row=row, column=item_col_index).value
            cost       = ws.cell(row=row, column=cost_col_index).value or 0
            qty_onhand = ws.cell(row=row, column=onhand_index).value or 0
            rlrctd     = ws.cell(row=row, column=rlrctd_index).value
            roll       = ws.cell(row=row, column=roll_col_index).value

            if sheet_name.strip().upper() == "AVERAGE COSTED":
                # For "Average Costed", ignore RROLL# in grouping → use empty string
                group_key = (warehouse, item, "")
            else:
                group_key = (warehouse, item, roll)

            initial_group[group_key].append({
                "row": row,
                "cost": cost,
                "qty":  qty_onhand,
                "rlrctd": rlrctd,
                "roll": roll
            })

        # === STEP 3: Compute LOT# per‐row, then regroup by (warehouse, item, LOT#) ===
        lot_group_map = defaultdict(list)
        row_to_lot    = {}  # maps row_idx → its LOT# string

        for (warehouse, item, roll), entries in initial_group.items():
            # Collect all RLRCTD values in this (warehouse, item, roll) block
            rlrctds = [e["rlrctd"] for e in entries]

            # Make a set of normalized date‐strings
            distinct_dates = set()
            for dt in rlrctds:
                if hasattr(dt, "strftime"):
                    distinct_dates.add(dt.strftime("%Y-%m-%d"))
                else:
                    distinct_dates.add(str(dt).strip())

            multiple_dates = (len(distinct_dates) > 1)

            for e in entries:
                row_idx = e["row"]
                if sheet_name.strip().upper() == "AVERAGE COSTED":
                    lot_str = ""  # always blank on this sheet
                else:
                    # Roll‐type sheets: build LOT# = "RROLL#" or "RROLL#-RLRCTD"
                    roll_str = str(e["roll"]).strip()
                    dt = e["rlrctd"]
                    if hasattr(dt, "strftime"):
                        dt_str = dt.strftime("%Y-%m-%d")
                    else:
                        dt_str = str(dt).strip().split(" ")[0]

                    if multiple_dates:
                        lot_str = f"{roll_str}-{dt_str}"
                    else:
                        lot_str = roll_str

                row_to_lot[row_idx] = lot_str
                lot_group_map[(warehouse, item, lot_str)].append({
                    "row": row_idx,
                    "cost": e["cost"],
                    "qty":  e["qty"]
                })

        # === STEP 4: Compute highestCost, averageCost, weightedCost per (warehouse, item, LOT#) ===
        computed_values = {}  # row_idx → {"lot":…, "highest":…, "average":…, "weighted":…}
        for (warehouse, item, lot_str), entries in lot_group_map.items():
            costs = [e["cost"] for e in entries]
            qtys  = [e["qty"]  for e in entries]

            highest = max(costs) if costs else 0
            average = (sum(costs) / len(costs)) if costs else 0

            total_qty = sum(qtys)
            if total_qty > 0:
                weighted = sum(c * q for c, q in zip(costs, qtys)) / total_qty
            else:
                weighted = 0

            # Assign to every row in this (warehouse, item, LOT#) subgroup
            for e in entries:
                computed_values[e["row"]] = {
                    "lot": lot_str,
                    "highest": highest,
                    "average": average,
                    "weighted": weighted
                }

        # === STEP 5: Insert four new headers (LOT#, highestCost, averageCost, weightedCost) ===
        last_col = ws.max_column
        lot_idx      = last_col + 1
        highest_idx  = last_col + 2
        average_idx  = last_col + 3
        weighted_idx = last_col + 4

        ws.cell(row=1, column=lot_idx,      value="LOT#")
        ws.cell(row=1, column=highest_idx,  value="highestCost")
        ws.cell(row=1, column=average_idx,  value="averageCost")
        ws.cell(row=1, column=weighted_idx, value="weightedCost")

        # === STEP 6: Populate those four columns with computed_values (numeric) ===
        for row_idx, vals in computed_values.items():
            ws.cell(row=row_idx, column=lot_idx,      value=vals["lot"])
            ws.cell(row=row_idx, column=highest_idx,  value=vals["highest"])
            ws.cell(row=row_idx, column=average_idx,  value=vals["average"])
            ws.cell(row=row_idx, column=weighted_idx, value=vals["weighted"])

        # === STEP 7: Insert seven new headers for Totals & Diffs ===
        # At this point, ws.max_column == last_col + 4, so the next indices are:
        total_cost_idx     = weighted_idx + 1  # last_col + 5
        total_highest_idx  = weighted_idx + 2  # last_col + 6
        total_average_idx  = weighted_idx + 3  # last_col + 7
        total_weighted_idx = weighted_idx + 4  # last_col + 8
        diff_highest_idx   = weighted_idx + 5  # last_col + 9
        diff_average_idx   = weighted_idx + 6  # last_col + 10
        diff_weighted_idx  = weighted_idx + 7  # last_col + 11

        ws.cell(row=1, column=total_cost_idx,     value="Total Cost")
        ws.cell(row=1, column=total_highest_idx,  value="Total Highest Cost")
        ws.cell(row=1, column=total_average_idx,  value="Total Average Cost")
        ws.cell(row=1, column=total_weighted_idx, value="Total Weighted Cost")
        ws.cell(row=1, column=diff_highest_idx,   value="Diff Highest")
        ws.cell(row=1, column=diff_average_idx,   value="Diff Average")
        ws.cell(row=1, column=diff_weighted_idx,  value="Diff Weighted")

        # === STEP 8: Populate those seven columns with Excel formulas ===
        # We need column letters for RONHAN, RLASTC, highestCost, averageCost, weightedCost, and the new totals.
        onhand_letter   = get_column_letter(onhand_index)
        cost_letter     = get_column_letter(cost_col_index)
        highest_letter  = get_column_letter(highest_idx)
        average_letter  = get_column_letter(average_idx)
        weighted_letter = get_column_letter(weighted_idx)
        total_cost_letter     = get_column_letter(total_cost_idx)
        total_highest_letter  = get_column_letter(total_highest_idx)
        total_average_letter  = get_column_letter(total_average_idx)
        total_weighted_letter = get_column_letter(total_weighted_idx)

        for row in range(2, ws.max_row + 1):
            # Build references to each cell in this row:
            onhand_cell   = f"${onhand_letter}${row}"
            cost_cell     = f"${cost_letter}${row}"
            highest_cell  = f"${highest_letter}${row}"
            average_cell  = f"${average_letter}${row}"
            weighted_cell = f"${weighted_letter}${row}"

            # --- Total Cost = RONHAN * RLASTC ---
            ws.cell(row=row, column=total_cost_idx).value = f"={onhand_cell} * {cost_cell}"

            # --- Total Highest Cost = RONHAN * highestCost ---
            ws.cell(row=row, column=total_highest_idx).value = f"={onhand_cell} * {highest_cell}"

            # --- Total Average Cost = RONHAN * averageCost ---
            ws.cell(row=row, column=total_average_idx).value = f"={onhand_cell} * {average_cell}"

            # --- Total Weighted Cost = RONHAN * weightedCost ---
            ws.cell(row=row, column=total_weighted_idx).value = f"={onhand_cell} * {weighted_cell}"

            # Now references to the new “Total” cells for this row:
            total_cost_cell     = f"${total_cost_letter}${row}"
            total_highest_cell  = f"${total_highest_letter}${row}"
            total_average_cell  = f"${total_average_letter}${row}"
            total_weighted_cell = f"${total_weighted_letter}${row}"

            # --- Diff Highest = Total Cost - Total Highest Cost ---
            ws.cell(row=row, column=diff_highest_idx).value = f"={total_cost_cell} - {total_highest_cell}"

            # --- Diff Average = Total Cost - Total Average Cost ---
            ws.cell(row=row, column=diff_average_idx).value = f"={total_cost_cell} - {total_average_cell}"

            # --- Diff Weighted = Total Cost - Total Weighted Cost ---
            ws.cell(row=row, column=diff_weighted_idx).value = f"={total_cost_cell} - {total_weighted_cell}"

        # === STEP 9: Autosize ALL columns (old + new) ===
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        # === STEP 10: Shade mismatched‐cost groups exactly as before ===
        shade_group_map = defaultdict(list)
        for row in range(2, ws.max_row + 1):
            warehouse = ws.cell(row=row, column=ware_col_index).value
            item      = ws.cell(row=row, column=item_col_index).value
            roll      = ws.cell(row=row, column=roll_col_index).value
            cost      = ws.cell(row=row, column=cost_col_index).value or 0
            key = (warehouse, item, roll)
            shade_group_map[key].append(cost)

        mismatched_groups = {
            key for key, cost_list in shade_group_map.items()
            if len({c for c in cost_list}) > 1
        }

        last_key = None
        fill_toggle = False

        for row in range(2, ws.max_row + 1):
            warehouse = ws.cell(row=row, column=ware_col_index).value
            item      = ws.cell(row=row, column=item_col_index).value
            roll      = ws.cell(row=row, column=roll_col_index).value
            group_key = (warehouse, item, roll)

            if group_key != last_key:
                fill_toggle = not fill_toggle
                last_key = group_key

            if group_key in mismatched_groups:
                row_fill = fill_red
            elif fill_toggle:
                row_fill = fill_gray
            else:
                row_fill = None

            if row_fill:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col_idx).fill = row_fill

        # Freeze the top row
        ws.freeze_panes = 'A2'

    # === STEP 11: Save workbook back to disk ===
    wb.save(file_path)
    print(f"✅ Shading complete (with Totals & Diffs added) in: {file_path}")


def main():
    suffix = input("Enter the suffix used for the classified inventory file: ").strip().replace(" ", "_") or "classified"
    path_manager = PathManager()
    file_path = path_manager.get_path("PATHS", "inventory_classified_path", suffix=suffix, check_path=False)
    autosize_and_shade_roll_groups(file_path)

if __name__ == "__main__":
    main()

