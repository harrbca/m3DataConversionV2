import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from path_manager import PathManager
from collections import defaultdict

def autosize_and_shade_roll_groups(file_path):
    fill_gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # === Autosize columns ===
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        # === Locate required columns ===
        headers = {str(cell.value).strip().upper(): i + 1 for i, cell in enumerate(ws[1])}
        roll_col_index = headers.get("RROLL#")
        cost_col_index = headers.get("RLASTC")
        ware_col_index = headers.get("RWARE#")
        item_col_index = headers.get("ITEMNUMBER") or headers.get("ITEM#")  # fallback for alt headers

        if not all([roll_col_index, cost_col_index, ware_col_index, item_col_index]):
            print(f"⚠️ Skipping sheet '{sheet_name}': missing required columns.")
            continue

        # === Build group_key → [(row#, cost)] map ===
        group_map = defaultdict(list)
        for row in range(2, ws.max_row + 1):
            roll = ws.cell(row=row, column=roll_col_index).value
            cost = ws.cell(row=row, column=cost_col_index).value
            warehouse = ws.cell(row=row, column=ware_col_index).value
            item = ws.cell(row=row, column=item_col_index).value
            group_key = (warehouse, item, roll)
            group_map[group_key].append((row, cost))

        # === Identify groups with mismatched costs ===
        mismatched_groups = {
            key for key, values in group_map.items()
            if len({c for _, c in values}) > 1
        }

        # === Apply alternating fill and red highlight ===
        last_key = None
        fill_toggle = False
        for row in range(2, ws.max_row + 1):
            warehouse = ws.cell(row=row, column=ware_col_index).value
            item = ws.cell(row=row, column=item_col_index).value
            roll = ws.cell(row=row, column=roll_col_index).value
            group_key = (warehouse, item, roll)

            if group_key != last_key:
                fill_toggle = not fill_toggle
                last_key = group_key

            if group_key in mismatched_groups:
                row_fill = fill_red
            elif fill_toggle:
                row_fill = fill_gray
            else:
                row_fill = None

            if row_fill:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = row_fill

        ws.freeze_panes = 'A2'

    wb.save(file_path)
    print(f"✅ Shading complete with cost mismatch detection in: {file_path}")


def main():
    suffix = input("Enter the suffix used for the classified inventory file: ").strip().replace(" ", "_") or "classified"
    path_manager = PathManager()
    file_path = path_manager.get_path("PATHS", "inventory_classified_path", suffix=suffix, check_path=False)

    autosize_and_shade_roll_groups(file_path)


if __name__ == "__main__":
    main()
