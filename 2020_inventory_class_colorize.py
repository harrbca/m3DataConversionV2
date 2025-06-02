import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from path_manager import PathManager

def autosize_and_shade_roll_groups(file_path):
    fill_gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # === Autosize columns ===
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        # === Find RROLL# column ===
        roll_col_index = next((i for i, cell in enumerate(ws[1], 1) if str(cell.value).strip().upper() == 'RROLL#'), None)
        if not roll_col_index:
            print(f"⚠️ Skipping sheet '{sheet_name}': 'RROLL#' column not found.")
            continue

        # === Apply shading on change of RROLL# ===
        last_roll = None
        fill_toggle = False

        for row in range(2, ws.max_row + 1):
            current_roll = ws.cell(row=row, column=roll_col_index).value

            if current_roll != last_roll:
                fill_toggle = not fill_toggle
                last_roll = current_roll

            if fill_toggle:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill_gray

        ws.freeze_panes = 'A2'

    wb.save(file_path)
    print(f"✅ Shading applied to RROLL# groups in: {file_path}")



def main():
    suffix = input("Enter the suffix used for the classified inventory file: ").strip().replace(" ", "_") or "classified"
    path_manager = PathManager()
    file_path = path_manager.get_path("PATHS", "inventory_classified_path", suffix=suffix, check_path=False)

    autosize_and_shade_roll_groups(file_path)


if __name__ == "__main__":
    main()
