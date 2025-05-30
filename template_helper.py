import os
import pandas as pd
from config_reader import ConfigReader
from path_manager import PathManager
from pathlib import Path
from openpyxl import load_workbook

class TemplateHelper:
    def __init__(self, template_name):
        config = ConfigReader.get_instance()
        template_base_path = config.get("PATHS", "template_path")
        template_path = Path(template_base_path) / template_name
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file {template_path} not found")

        print(f"Loading template from {template_path}")
        self.wb = load_workbook(template_path)
        self.ws = self.wb.active
        self.column_map = self._map_columns()
        self.data_start_row = 4

    def _map_columns(self):
        # the first row contains column names
        column_map = {}
        for col in range(1, self.ws.max_column + 1):
            column_name = self.ws.cell(row=1, column=col).value
            cell_name = self.ws.cell(row=1, column=col).column_letter
            column_map[column_name] = cell_name

        return column_map

    def add_row(self, data):
        row = self.ws.max_row + 1
        for column_name, value in data.items():
            if column_name not in self.column_map:
                raise ValueError(f"Column name {column_name} not found in template")
            self.ws[f"{self.column_map[column_name]}{row}"] = value

    def add_all_rows(self, entries: list[dict]):
        start_row = self.ws.max_row + 1
        for i, data in enumerate(entries):
            row_num = start_row + i
            for column_name, value in data.items():
                if column_name not in self.column_map:
                    raise ValueError(f"Column name {column_name} not found in template")
                self.ws[f"{self.column_map[column_name]}{row_num}"] = value


    def save(self, output_file):
        path_manager = PathManager()
        output_path = path_manager.get_path("PATHS", output_file)
        print(f"Saving to {output_path}")
        self.wb.save(output_path)