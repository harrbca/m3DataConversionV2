import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from path_manager import PathManager
from collections import defaultdict

def autosize_and_shade_roll_groups(file_path):
    fill_gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    fill_red  = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # === STEP 1: Header → column‐index map (uppercase keys) ===
        headers = {
            str(cell.value).strip().upper(): idx + 1
            for idx, cell in enumerate(ws[1])
        }

        # Required for shading logic:
        roll_col_index  = headers.get("RROLL#")
        cost_col_index  = headers.get("RLASTC")
        ware_col_index  = headers.get("RWARE#")
        item_col_index  = headers.get("ITEMNUMBER") or headers.get("ITEM#")
        # New columns needed for LOT# + weighted cost:
        rlrctd_index    = headers.get("RLRCTD")
        onhand_index    = headers.get("RONHAN")

        if not all([ware_col_index, item_col_index, cost_col_index, rlrctd_index, onhand_index]):
            print(f"⚠️ Skipping sheet '{sheet_name}': missing one of RWARE#, ITEMNUMBER/ITEM#, RLASTC, RLRCTD, or RONHAN.")
            continue

        # === STEP 2: Build an initial group_map by (warehouse, item, roll)
        # We’ll use this first to decide each row’s LOT#, then regroup by (warehouse, item, LOT#).
        initial_group = defaultdict(list)
        for row in range(2, ws.max_row + 1):
            warehouse = ws.cell(row=row, column=ware_col_index).value
            item      = ws.cell(row=row, column=item_col_index).value
            cost      = ws.cell(row=row, column=cost_col_index).value or 0
            qty_onhand= ws.cell(row=row, column=onhand_index).value or 0
            rlrctd    = ws.cell(row=row, column=rlrctd_index).value
            roll      = ws.cell(row=row, column=roll_col_index).value

            if sheet_name.strip().upper() == "AVERAGE COSTED":
                # For Average Costed: we do not care about 'roll' in grouping → use empty string
                group_key = (warehouse, item, "")
            else:
                group_key = (warehouse, item, roll)

            initial_group[group_key].append({
                "row": row,
                "cost": cost,
                "qty":  qty_onhand,
                "rlrctd": rlrctd,
                "roll": roll
            })

        # === STEP 3: For each (warehouse, item, roll) group, assign LOT# per‐row ===
        # Then accumulate each row into lot_group_map keyed by (warehouse, item, lot).
        lot_group_map = defaultdict(list)
        row_to_lot    = {}  # temporary mapping of row → its lot string

        for (warehouse, item, roll), entries in initial_group.items():
            # Extract all RLRCTD values in this (warehouse, item, roll) block
            rlrctds = [e["rlrctd"] for e in entries]

            # Determine if there are multiple distinct RLRCTD within this block
            distinct_dates = set()
            for dt in rlrctds:
                # Normalize date to string "YYYY-MM-DD" if datetime, else strip whitespace
                if hasattr(dt, "strftime"):
                    distinct_dates.add(dt.strftime("%Y-%m-%d"))
                else:
                    distinct_dates.add(str(dt).strip())

            multiple_dates = (len(distinct_dates) > 1)

            for e in entries:
                row_idx = e["row"]
                # Build LOT# depending on multiple_dates or not
                if sheet_name.strip().upper() == "AVERAGE COSTED":
                    lot_str = ""  # always blank
                else:
                    # Roll‐type: strip whitespace from both roll and rlrctd, then concatenate if needed
                    roll_str = str(e["roll"]).strip()
                    # Normalize this row's RLRCTD
                    dt = e["rlrctd"]
                    if hasattr(dt, "strftime"):
                        dt_str = dt.strftime("%Y-%m-%d")
                    else:
                        dt_str = str(dt).strip().split(" ")[0]

                    if multiple_dates:
                        lot_str = f"{roll_str}-{dt_str}"
                    else:
                        lot_str = roll_str

                # Remember each row's LOT#
                row_to_lot[row_idx] = lot_str

                # Accumulate into lot_group_map: key = (warehouse, item, lot_str)
                lot_group_map[(warehouse, item, lot_str)].append({
                    "row": row_idx,
                    "cost": e["cost"],
                    "qty":  e["qty"]
                })

        # === STEP 4: Compute highestCost, averageCost, weightedCost per (warehouse, item, lot) ===
        computed_values = {}  # row_idx → dict(lot, highest, average, weighted)
        for (warehouse, item, lot_str), entries in lot_group_map.items():
            costs = [e["cost"] for e in entries]
            qtys  = [e["qty"]  for e in entries]

            highest = max(costs) if costs else 0
            average = (sum(costs) / len(costs)) if costs else 0

            total_qty = sum(qtys)
            if total_qty > 0:
                weighted = sum(c * q for c, q in zip(costs, qtys)) / total_qty
            else:
                weighted = 0

            # Assign these four values to every row in this lot‐group
            for e in entries:
                computed_values[e["row"]] = {
                    "lot": lot_str,
                    "highest": highest,
                    "average": average,
                    "weighted": weighted
                }

        # === STEP 5: Insert four new headers at far right of row 1 ===
        last_col = ws.max_column
        ws.cell(row=1, column= last_col + 1, value="LOT#")
        ws.cell(row=1, column= last_col + 2, value="highestCost")
        ws.cell(row=1, column= last_col + 3, value="averageCost")
        ws.cell(row=1, column= last_col + 4, value="weightedCost")

        # === STEP 6: Populate those columns using computed_values ===
        for row_idx, vals in computed_values.items():
            ws.cell(row=row_idx, column= last_col + 1, value= vals["lot"])
            ws.cell(row=row_idx, column= last_col + 2, value= vals["highest"])
            ws.cell(row=row_idx, column= last_col + 3, value= vals["average"])
            ws.cell(row=row_idx, column= last_col + 4, value= vals["weighted"])

        # === STEP 7: Autosize ALL columns (old + new) ===
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        # === STEP 8: Shade mismatched‐cost groups exactly as before ===
        # Build a fresh shade_group_map keyed by (warehouse, item, roll) using original cost
        shade_group_map = defaultdict(list)
        for row in range(2, ws.max_row + 1):
            warehouse = ws.cell(row=row, column=ware_col_index).value
            item      = ws.cell(row=row, column=item_col_index).value
            roll      = ws.cell(row=row, column=roll_col_index).value
            cost      = ws.cell(row=row, column=cost_col_index).value or 0
            key = (warehouse, item, roll)
            shade_group_map[key].append(cost)

        mismatched_groups = {
            key for key, cost_list in shade_group_map.items()
            if len({c for c in cost_list}) > 1
        }

        # Now shade every row, toggling gray for each new (warehouse, item, roll) block,
        # or red if its (warehouse, item, roll) is in mismatched_groups
        last_key = None
        fill_toggle = False

        for row in range(2, ws.max_row + 1):
            warehouse = ws.cell(row=row, column=ware_col_index).value
            item      = ws.cell(row=row, column=item_col_index).value
            roll      = ws.cell(row=row, column=roll_col_index).value
            group_key = (warehouse, item, roll)

            if group_key != last_key:
                fill_toggle = not fill_toggle
                last_key = group_key

            if group_key in mismatched_groups:
                row_fill = fill_red
            elif fill_toggle:
                row_fill = fill_gray
            else:
                row_fill = None

            if row_fill:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = row_fill

        # Freeze the top row
        ws.freeze_panes = 'A2'

    # === STEP 9: Save workbook back to disk ===
    wb.save(file_path)
    print(f"✅ Shading complete (with corrected LOT# grouping and cost‐calculations) in: {file_path}")


def main():
    suffix = input("Enter the suffix used for the classified inventory file: ").strip().replace(" ", "_") or "classified"
    path_manager = PathManager()
    file_path = path_manager.get_path("PATHS", "inventory_classified_path", suffix=suffix, check_path=False)
    autosize_and_shade_roll_groups(file_path)

if __name__ == "__main__":
    main()


