from openpyxl import load_workbook

class ExcelTemplate:
    def __init__(self, template_file):
        """
        Initializes the ExcelTemplate object.

        :param template_file: Path to the Excel template file.
        """
        self.template_file = template_file
        self.wb = load_workbook(template_file)
        self.ws = self.wb.active  # Modify if a specific sheet is needed
        self.column_mapping = self._load_column_mapping()

    def _load_column_mapping(self):
        """
                Reads column headers from row 1 and maps them to column indexes.

                :return: Dictionary mapping column names to column numbers.
                """
        mapping = {}
        for col in range(1, self.ws.max_column + 1):  # Loop through columns
            column_name = self.ws.cell(row=1, column=col).value  # Get column header
            if column_name:
                mapping[column_name] = col  # Store column index by name
        return mapping

    def set_value(self, row, column_name, value):
        """
        Sets a value in the Excel sheet at the specified row and column.

        :param row: The row number where the value should be inserted (1-based index).
        :param column_name: The column name as defined in row 1 of the Excel template.
        :param value: The value to insert.
        """
        if column_name in self.column_mapping:  # Ensure column exists
            col_index = self.column_mapping[column_name]
            self.ws.cell(row=row, column=col_index, value=value)
        else:
            raise ValueError(f"Column '{column_name}' not found in Excel template.")

    def save(self, output_file):
        """
        Saves the modified Excel file.

        :param output_file: Path where the updated file will be saved.
        """
        self.wb.save(output_file)

    def adjust_column_widths(self):
        """
        Adjusts column widths based on the length of the longest value in each column.
        """
        try:
            for col in self.ws.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Get the column letter (A, B, C, etc.)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2  # Add some padding
                self.ws.column_dimensions[col_letter].width = adjusted_width

            print(f"✅ Column widths adjusted in {self.template_file}")

        except Exception as e:
            print(f"❌ Error adjusting column widths: {e}")
